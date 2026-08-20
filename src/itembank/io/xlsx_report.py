"""教員用照合表と統計レポートの xlsx 出力(設計書 §13.2)。

集計は ``core.reporting`` が済ませてあり、ここは並べて書くだけ。
xlsx 出力ではタグを除去して用いる(設計書 §3.2)。
"""

from __future__ import annotations

from collections.abc import Sequence
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

from ..core.reporting import (
    CrosswalkRow,
    ReportRow,
    Stratum,
    stratify_by_negative,
    stratify_by_type,
)
from ..core.text import strip_tags

HEADER_FONT = Font(bold=True)
PERCENT = "0.0%"
DECIMAL3 = "0.000"


def _write_header(sheet: Worksheet, headers: Sequence[str]) -> None:
    sheet.append(list(headers))
    for cell in sheet[1]:
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center")
    sheet.freeze_panes = "A2"


def _autosize(sheet: Worksheet, *, max_width: int = 48) -> None:
    for column in sheet.columns:
        width = max((len(str(c.value)) for c in column if c.value is not None), default=0)
        sheet.column_dimensions[get_column_letter(column[0].column)].width = min(
            max(10, width + 2), max_width
        )


def _yesno(value: bool) -> str:
    return "○" if value else ""


# ---------------------------------------------------------------------------
# 教員用照合表(設計書 §13.2)
# ---------------------------------------------------------------------------

CROSSWALK_HEADERS = (
    "出題番号",
    "問題ID",
    "版",
    "タイプ",
    "否定形",
    "新作",
    "正答",
    "分野",
    "前回出題年",
    "前回正答率",
    "使用セットID",
)


def write_crosswalk(
    rows: Sequence[CrosswalkRow], path: Path | str, *, exam_name: str | None = None
) -> Path:
    """出題番号・問題ID・版・タイプ・否定形の別・新作の別・正答・分野・
    前回出題年・前回正答率・使用セットID を並べる。"""
    book = Workbook()
    sheet = book.active
    sheet.title = "照合表"
    _write_header(sheet, CROSSWALK_HEADERS)

    for row in sorted(rows, key=lambda r: r.position):
        sheet.append(
            [
                row.position,
                row.question_id,
                row.version_no,
                row.item_type or "?",
                _yesno(row.negative),
                _yesno(row.is_new),
                row.correct,
                row.tags,
                row.last_exam_year,
                row.last_p,
                row.choice_set_id,
            ]
        )
    for cell in sheet["J"][1:]:
        cell.number_format = PERCENT

    if exam_name:
        sheet.insert_rows(1)
        sheet["A1"] = f"教員用照合表: {exam_name}"
        sheet["A1"].font = Font(bold=True, size=12)
        sheet.freeze_panes = "A3"

    _autosize(sheet)
    out = Path(path)
    out.parent.mkdir(parents=True, exist_ok=True)
    book.save(str(out))
    return out


# ---------------------------------------------------------------------------
# 統計レポート(設計書 §13.2)
# ---------------------------------------------------------------------------

ITEM_HEADERS = (
    "出題番号",
    "問題ID",
    "版",
    "タイプ",
    "否定形",
    "正答",
    "分野",
    "N",
    "正答数",
    "正答率",
    "正答率(タイプ併記)",
    "識別係数",
    "a",
    "b",
    "c",
    "d",
    "e",
    "無回答率",
    "指示個数違反率",
    "最頻誤答",
    "最頻誤答度数",
    "前回正答率",
    "前回比",
    "フラグ",
)

STRATUM_HEADERS = ("区分", "問題数", "平均正答率", "中央正答率", "平均識別係数")


def write_stats_report(
    rows: Sequence[ReportRow],
    path: Path | str,
    *,
    exam_name: str | None = None,
    stem_texts: dict[int, str] | None = None,
) -> Path:
    """タイプ別の問題別一覧・否定形/肯定形の層別集計・パターン分析・前回比較。"""
    book = Workbook()
    _sheet_items(book.active, rows, stem_texts or {})
    _sheet_strata(book.create_sheet("層別集計"), rows)
    _sheet_patterns(book.create_sheet("パターン分析"), rows)
    _sheet_comparison(book.create_sheet("前回比較"), rows)
    _sheet_flags(book.create_sheet("フラグ"), rows)

    if exam_name:
        book.properties.title = f"統計レポート: {exam_name}"

    out = Path(path)
    out.parent.mkdir(parents=True, exist_ok=True)
    book.save(str(out))
    return out


#: 一覧はタイプ別にまとめる(設計書 §13.1 の候補一覧と同じ並び)。
_TYPE_ORDER = {"A": 0, "X2": 1, "X3": 2, "X4": 3, "XX": 4}


def _by_type_then_position(row: ReportRow) -> tuple[int, int]:
    return (_TYPE_ORDER.get(row.item_type or "", 9), row.position)


def _sheet_items(sheet: Worksheet, rows: Sequence[ReportRow], stem_texts: dict[int, str]) -> None:
    sheet.title = "問題別一覧"
    headers = (*ITEM_HEADERS, "設問文") if stem_texts else ITEM_HEADERS
    _write_header(sheet, headers)

    for row in sorted(rows, key=_by_type_then_position):
        values = [
            row.position,
            row.question_id,
            row.version_no,
            row.item_type or "?",
            _yesno(row.negative),
            row.correct,
            row.tags,
            row.n,
            row.n_correct,
            row.p,
            row.p_label,
            row.disc,
            *[row.sel.get(label) for label in "abcde"],
            row.blank_rate,
            row.overselect_rate,
            row.top_wrong_pattern,
            row.top_wrong_count,
            row.prev_p,
            row.delta_p,
            ", ".join(row.flags),
        ]
        if stem_texts:
            # xlsx 出力ではタグを除去して用いる(設計書 §3.2)。
            values.append(strip_tags(stem_texts.get(row.position, "")))
        sheet.append(values)

    for letter in ("J", "M", "N", "O", "P", "Q", "R", "S", "V", "W"):
        for cell in sheet[letter][1:]:
            cell.number_format = PERCENT
    for cell in sheet["L"][1:]:
        cell.number_format = DECIMAL3
    _autosize(sheet)


def _append_strata(sheet: Worksheet, title: str, strata: Sequence[Stratum]) -> None:
    sheet.append([title])
    sheet.cell(row=sheet.max_row, column=1).font = HEADER_FONT
    for s in strata:
        sheet.append([s.name, s.n_items, s.mean_p, s.median_p, s.mean_disc])
    sheet.append([])


def _sheet_strata(sheet: Worksheet, rows: Sequence[ReportRow]) -> None:
    """否定形/肯定形の層別集計(設計書 §4-(2))。

    「この問題が難しいのは内容のせいか、否定形だからか」を切り分ける材料。
    """
    _write_header(sheet, STRATUM_HEADERS)
    _append_strata(sheet, "否定形/肯定形", stratify_by_negative(rows))
    _append_strata(sheet, "タイプ別", stratify_by_type(rows))

    for letter in ("C", "D"):
        for cell in sheet[letter][1:]:
            cell.number_format = PERCENT
    for cell in sheet["E"][1:]:
        cell.number_format = DECIMAL3
    _autosize(sheet)


def _sheet_patterns(sheet: Worksheet, rows: Sequence[ReportRow]) -> None:
    """誤答パターン上位 5 件(設計書 §14-3)。"""
    _write_header(sheet, ("出題番号", "タイプ", "正答", "順位", "誤答パターン", "度数", "割合"))
    for row in sorted(rows, key=lambda r: r.position):
        for rank, (pattern, count) in enumerate(row.top_wrong, start=1):
            sheet.append(
                [
                    row.position,
                    row.item_type or "?",
                    row.correct,
                    rank,
                    pattern,
                    count,
                    (count / row.n) if row.n else None,
                ]
            )
    for cell in sheet["G"][1:]:
        cell.number_format = PERCENT
    _autosize(sheet)


def _sheet_comparison(sheet: Worksheet, rows: Sequence[ReportRow]) -> None:
    _write_header(
        sheet, ("出題番号", "問題ID", "タイプ", "今回正答率", "前回正答率", "差", "今回D", "前回D")
    )
    for row in sorted(rows, key=lambda r: r.position):
        if row.prev_p is None and row.prev_disc is None:
            continue
        sheet.append(
            [
                row.position,
                row.question_id,
                row.item_type or "?",
                row.p,
                row.prev_p,
                row.delta_p,
                row.disc,
                row.prev_disc,
            ]
        )
    for letter in ("D", "E", "F"):
        for cell in sheet[letter][1:]:
            cell.number_format = PERCENT
    for letter in ("G", "H"):
        for cell in sheet[letter][1:]:
            cell.number_format = DECIMAL3
    _autosize(sheet)


def _sheet_flags(sheet: Worksheet, rows: Sequence[ReportRow]) -> None:
    """取込後の振り返り用。ここから改訂に進む(設計書 §9.3, §2.6)。"""
    _write_header(sheet, ("出題番号", "問題ID", "タイプ", "正答率", "識別係数", "フラグ"))
    for row in sorted(rows, key=lambda r: r.position):
        if not row.flags:
            continue
        sheet.append(
            [
                row.position,
                row.question_id,
                row.item_type or "?",
                row.p,
                row.disc,
                ", ".join(row.flags),
            ]
        )
    for cell in sheet["D"][1:]:
        cell.number_format = PERCENT
    for cell in sheet["E"][1:]:
        cell.number_format = DECIMAL3
    _autosize(sheet)
