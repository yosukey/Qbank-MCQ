"""レポートの集計と xlsx 出力(設計書 §4-(2), §6.5, §13.2)。"""

from __future__ import annotations

from dataclasses import dataclass
from pathlib import Path

import pytest
from openpyxl import load_workbook
from sqlalchemy.orm import Session

from qbank_mcq.core.bank import create_question_from_printed
from qbank_mcq.core.exam import apply_stats, create_exam, finalize_exam, set_exam_items
from qbank_mcq.core.reporting import (
    ReportRow,
    choice_item_appearances,
    crosswalk_rows,
    report_rows,
    stratify_by_negative,
    stratify_by_type,
)
from qbank_mcq.core.stats import aggregate_item_performance
from qbank_mcq.io.xlsx_report import write_crosswalk, write_stats_report

SET_A = ["エナメル質", "象牙質", "セメント質", "歯髄", "歯根膜"]
SET_B = ["レッチウス条", "エブネル線", "新産線", "トームス突起", "シュレーゲル条"]

POSITIVE = "最も硬いのはどれか。1つ選べ。"
NEGATIVE = "石灰化し<strong>ない</strong>のはどれか。1つ選べ。"


@dataclass
class Row:
    position: int
    correct: str
    _counts: dict[str, int]
    disc: float | None = None

    def counts(self) -> dict[str, int]:
        return dict(self._counts)


def row(position: int, *, p: float | None, negative: bool = False, **kw) -> ReportRow:
    defaults = dict(
        question_id=position,
        version_no=1,
        item_type="A",
        negative=negative,
        correct="a",
        tags="",
        n=100,
        n_correct=int((p or 0) * 100),
        p=p,
        disc=0.3,
        sel={c: 0.2 for c in "abcde"},
        blank_rate=0.0,
        overselect_rate=0.0,
        top_wrong_pattern="b",
        top_wrong_count=10,
        flags=[],
    )
    defaults.update(kw)
    return ReportRow(position=position, **defaults)


# ---------------------------------------------------------------------------
# 層別集計(設計書 §4-(2))
# ---------------------------------------------------------------------------


def test_negative_and_positive_are_stratified_separately() -> None:
    """「この問題が難しいのは内容のせいか、否定形だからか」を切り分ける材料。"""
    rows = [
        row(1, p=0.9),
        row(2, p=0.8),
        row(3, p=0.4, negative=True),
        row(4, p=0.5, negative=True),
    ]
    strata = {s.name: s for s in stratify_by_negative(rows)}
    assert strata["否定形"].n_items == 2
    assert strata["肯定形"].n_items == 2
    assert strata["否定形"].mean_p == pytest.approx(0.45)
    assert strata["肯定形"].mean_p == pytest.approx(0.85)
    assert strata["全体"].n_items == 4


def test_type_strata_follow_the_documented_order() -> None:
    rows = [
        row(1, p=0.5, item_type="XX"),
        row(2, p=0.5, item_type="A"),
        row(3, p=0.5, item_type="X2"),
    ]
    assert [s.name for s in stratify_by_type(rows)] == ["A", "X2", "XX"]


def test_stratum_of_empty_rows_has_no_mean() -> None:
    strata = {s.name: s for s in stratify_by_negative([row(1, p=0.5)])}
    assert strata["否定形"].n_items == 0
    assert strata["否定形"].mean_p is None


def test_p_label_always_carries_the_type() -> None:
    """正答率は必ずタイプと併記する(設計書 §12)。"""
    assert row(1, p=0.41, item_type="X2").p_label == "41%(X2)"
    assert row(2, p=None).p_label == "—(A)"


def test_delta_p_compares_with_the_previous_exam() -> None:
    assert row(1, p=0.7, prev_p=0.5).delta_p == pytest.approx(0.2)
    assert row(2, p=0.7).delta_p is None


# ---------------------------------------------------------------------------
# DB からの組み立て
# ---------------------------------------------------------------------------


def add(session: Session, stem: str, correct: str, items):
    result, _ = create_question_from_printed(
        session, stem_html=stem, printed_choices=list(items), correct=correct
    )
    assert not result.blocked
    return result


@pytest.fixture
def imported_exam(session: Session):
    q1 = add(session, POSITIVE, "a", SET_A)
    q2 = add(session, NEGATIVE, "b", SET_B)
    exam = create_exam(session, name="定期試験2025", exam_date="2025-08-25")
    set_exam_items(session, exam, [(1, q1.version.id), (2, q2.version.id)])
    finalize_exam(session, exam)
    apply_stats(
        session,
        exam,
        [
            Row(1, "a", {"a": 90, "b": 6, "c": 2, "d": 1, "e": 1}, disc=0.4),
            Row(2, "b", {"b": 40, "a": 55, "c": 3, "d": 1, "e": 1}, disc=0.1),
        ],
        n_examinees=100,
        disc_type="D_25",
    )
    return exam


def test_report_rows_carry_the_derived_values(session: Session, imported_exam) -> None:
    rows = {r.position: r for r in report_rows(session, imported_exam)}
    assert rows[1].n == 100
    assert rows[1].p == pytest.approx(0.90)
    assert rows[1].negative is False
    assert rows[2].negative is True
    assert rows[2].top_wrong_pattern == "a"


def test_report_rows_list_the_top_wrong_patterns(session: Session, imported_exam) -> None:
    """誤答パターン上位 5 件(設計書 §14-3)。"""
    rows = {r.position: r for r in report_rows(session, imported_exam)}
    assert rows[2].top_wrong[0] == ("a", 55)
    assert len(rows[2].top_wrong) <= 5
    assert all(p != "b" for p, _ in rows[2].top_wrong)  # 正答は入らない


def test_dominant_wrong_is_flagged(session: Session, imported_exam) -> None:
    rows = {r.position: r for r in report_rows(session, imported_exam)}
    assert "dominant_wrong" in rows[2].flags


def test_crosswalk_rows_mark_new_items(session: Session, imported_exam) -> None:
    rows = crosswalk_rows(session, imported_exam)
    assert [r.position for r in rows] == [1, 2]
    assert all(r.is_new for r in rows)  # 前回統計が無いので新作扱い
    assert rows[1].negative is True
    assert rows[0].item_type == "A"


# ---------------------------------------------------------------------------
# xlsx 出力(設計書 §13.2)
# ---------------------------------------------------------------------------


def test_crosswalk_xlsx_has_the_documented_columns(
    session: Session, imported_exam, tmp_path: Path
) -> None:
    path = write_crosswalk(
        crosswalk_rows(session, imported_exam), tmp_path / "cw.xlsx", exam_name="定期試験2025"
    )
    sheet = load_workbook(path)["照合表"]
    assert sheet["A1"].value == "教員用照合表: 定期試験2025"
    headers = [c.value for c in sheet[2]]
    for expected in ("出題番号", "問題ID", "版", "タイプ", "否定形", "新作", "正答", "分野"):
        assert expected in headers
    assert "前回出題年" in headers and "使用セットID" in headers


def test_stats_report_has_every_documented_sheet(
    session: Session, imported_exam, tmp_path: Path
) -> None:
    rows = report_rows(session, imported_exam)
    path = write_stats_report(rows, tmp_path / "r.xlsx", exam_name="定期試験2025")
    book = load_workbook(path)
    assert book.sheetnames == ["問題別一覧", "層別集計", "パターン分析", "前回比較", "フラグ"]

    items = book["問題別一覧"]
    assert items.max_row == 3  # ヘッダ + 2 問
    strata = book["層別集計"]
    text = "\n".join(str(c.value) for r in strata.iter_rows() for c in r if c.value)
    assert "否定形" in text and "肯定形" in text and "タイプ別" in text


def test_report_strips_tags_from_stem_text(session: Session, imported_exam, tmp_path: Path) -> None:
    """xlsx 出力ではタグを除去して用いる(設計書 §3.2)。"""
    rows = report_rows(session, imported_exam)
    path = write_stats_report(rows, tmp_path / "r.xlsx", stem_texts={1: POSITIVE, 2: NEGATIVE})
    sheet = load_workbook(path)["問題別一覧"]
    stems = [row[-1] for row in sheet.iter_rows(min_row=2, values_only=True)]
    assert "石灰化しないのはどれか。1つ選べ。" in stems
    assert not any("<strong>" in str(s) for s in stems)


def test_pattern_sheet_lists_wrong_patterns(
    session: Session, imported_exam, tmp_path: Path
) -> None:
    path = write_stats_report(report_rows(session, imported_exam), tmp_path / "r.xlsx")
    sheet = load_workbook(path)["パターン分析"]
    rows = list(sheet.iter_rows(min_row=2, values_only=True))
    assert any(r[4] == "a" and r[5] == 55 for r in rows)


# ---------------------------------------------------------------------------
# 選択肢アイテム単位の実績(設計書 §6.5)
# ---------------------------------------------------------------------------


def test_item_performance_tracks_terms_across_questions(session: Session, imported_exam) -> None:
    appearances = choice_item_appearances(session)
    assert appearances

    performance = {p.text_html: p for p in aggregate_item_performance(appearances)}
    enamel = performance["エナメル質"]
    assert enamel.appearances == 1
    assert enamel.as_correct == 1
    assert enamel.median_p_when_correct == pytest.approx(0.90)

    dentin = performance["象牙質"]
    assert dentin.as_distractor == 1
    assert dentin.median_mark_rate_when_distractor is not None


def test_most_confused_partner_is_identified(session: Session, imported_exam) -> None:
    """「最も混同される相手」= 正答だったときに誤選択の主軸になった項目(設計書 §6.5)。"""
    performance = {
        p.text_html: p for p in aggregate_item_performance(choice_item_appearances(session))
    }
    # 問2 の正答は「エブネル線」で、55 名が「レッチウス条」に流れている。
    assert performance["エブネル線"].top_confused_with == "レッチウス条"
    assert performance["エブネル線"].top_confused_count == 1


def test_item_performance_is_sorted_by_appearances() -> None:
    from qbank_mcq.core.stats import ItemAppearance

    appearances = [
        ItemAppearance("まれな語", 1, 1, False, 0.1, None),
        ItemAppearance("よく出る語", 1, 1, False, 0.1, None),
        ItemAppearance("よく出る語", 2, 2, False, 0.2, None),
    ]
    result = aggregate_item_performance(appearances)
    assert [p.text_html for p in result] == ["よく出る語", "まれな語"]
